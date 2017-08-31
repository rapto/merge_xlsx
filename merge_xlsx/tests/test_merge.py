import unittest
import pkg_resources
import openpyxl
import datetime

from merge_xlsx import merge
from path import Path
from merge_xlsx.util import TemporaryDirectory

class TestXlsxMerge(unittest.TestCase):
    def te_st_merge(self, **merged_data):
        f = pkg_resources.resource_filename('merge_xlsx.tests', 'fixtures/template.xlsx')
        with TemporaryDirectory() as td:
            result = Path(td) / 'result.xlsx'
            merge(f, result, merged_data)
            xlsx = openpyxl.load_workbook(result)
            ws = xlsx.get_sheet_by_name('Hoja1')
            for k, v in merged_data.items(): 
                c = ws[k] 
                self.assertEqual(c.value, v)

    def test_merge_string(self):
        self.te_st_merge(B8=u'Marcos S\xe1nchez Provencio')

    def test_merge_date(self):
        self.te_st_merge(D7=datetime.datetime(2017,8,29))
    
    def test_merge_number(self):
        self.te_st_merge(C12=666)
    
    def test_merge_image(self):
        img = pkg_resources.resource_stream('merge_xlsx.tests', 'fixtures/img.png')
        f = pkg_resources.resource_filename('merge_xlsx.tests', 'fixtures/template.xlsx')
        merged_data = {'Image logo': img}
        with TemporaryDirectory() as td:
            result = Path(td) / 'result.xlsx'
            merge(f, result, merged_data)
            xlsx = openpyxl.load_workbook(result)
            # FIXME only test for errors, can't check content
            